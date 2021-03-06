# coding=utf-8
# Copyright 2018 The Google AI Language Team Authors.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
"""Run BERT on SQuAD 1.1 and SQuAD 2.0."""

from __future__ import absolute_import
from __future__ import division
from __future__ import print_function

import collections
import json
import math
import os, types
import random
import modeling
import optimization
import tokenization
import six
import copy
import tensorflow as tf
import numpy as np
import scipy.sparse as sp

# do excel
from openpyxl import Workbook


import uuid

# do 
import code
import prettytable

from decimal import *
import decimal
getcontext().prec = 50

#Willy Define
example_in_set_eval_examples = 0
example_in_write_predictions = 0
predict_result_index = 0
checkState_in_AtenResult = 0
checkState_in_AtenResult2 = 0
checkState_in_GetAnswer = 0
checkState_add_retriever = 0
willy_check_code = "willy test on 201907101548"

from drqa import retriever

DOC2IDX = None
documents = []
db_class = retriever.get_class('sqlite')



flags = tf.flags

FLAGS = flags.FLAGS

## Required parameters
flags.DEFINE_string(
    "bert_config_file", None,
    "The config json file corresponding to the pre-trained BERT model. "
    "This specifies the model architecture.")

flags.DEFINE_string("vocab_file", None,
                    "The vocabulary file that the BERT model was trained on.")

flags.DEFINE_string(
    "output_dir", None,
    "The output directory where the model checkpoints will be written.")

## Other parameters
flags.DEFINE_string("train_file", None,
                    "SQuAD json for training. E.g., train-v1.1.json")

flags.DEFINE_string(
    "predict_file", None,
    "SQuAD json for predictions. E.g., dev-v1.1.json or test-v1.1.json")

flags.DEFINE_string(
    "init_checkpoint", None,
    "Initial checkpoint (usually from a pre-trained BERT model).")

flags.DEFINE_bool(
    "do_lower_case", True,
    "Whether to lower case the input text. Should be True for uncased "
    "models and False for cased models.")

flags.DEFINE_integer(
    "max_seq_length", 384,
    "The maximum total input sequence length after WordPido_interactiveece tokenization. "
    "Sequences longer than this will be truncated, and sequences shorter "
    "than this will be padded.")

flags.DEFINE_integer(
    "doc_stride", 128,
    "When splitting up a long document into chunks, how much stride to "
    "take between chunks.")

flags.DEFINE_integer(
    "max_query_length", 64,
    "The maximum number of tokens for the question. Questions longer than "
    "this will be truncated to this length.")

flags.DEFINE_bool("do_train", False, "Whether to run training.")

flags.DEFINE_bool("do_predict", False, "Whether to run eval on the dev set.")

flags.DEFINE_integer("train_batch_size", 32, "Total batch size for training.")

flags.DEFINE_integer("predict_batch_size", 8,
                     "Total batch size for predictions.")

flags.DEFINE_float("learning_rate", 5e-5, "The initial learning rate for Adam.")

flags.DEFINE_float("num_train_epochs", 3.0,
                   "Total number of training epochs to perform.")

flags.DEFINE_float(
    "warmup_proportion", 0.1,
    "Proportion of training to perform linear learning rate warmup for. "
    "E.g., 0.1 = 10% of training.")

flags.DEFINE_integer("save_checkpoints_steps", 1000,
                     "How often to save the model checkpoint.")

flags.DEFINE_integer("iterations_per_loop", 1000,
                     "How many steps to make in each estimator call.")

flags.DEFINE_integer(
    "n_best_size", 20,
    "The total number of n-best predictions to generate in the "
    "nbest_predictions.json output file.")

flags.DEFINE_integer(
    "max_answer_length", 30,
    "The maximum length of an answer that can be generated. This is needed "
    "because the start and end predictions are not conditioned on one another.")

flags.DEFINE_bool("use_tpu", False, "Whether to use TPU or GPU/CPU.")

tf.flags.DEFINE_string(
    "tpu_name", None,
    "The Cloud TPU to use for training. This should be either the name "
    "used when creating the Cloud TPU, or a grpc://ip.address.of.tpu:8470 "
    "url.")

tf.flags.DEFINE_string(
    "tpu_zone", None,
    "[Optional] GCE zone where the Cloud TPU is located in. If not "
    "specified, we will attempt to automatically detect the GCE project from "
    "metadata.")

tf.flags.DEFINE_string(
    "gcp_project", None,
    "[Optional] Project name for the Cloud TPU-enabled project. If not "
    "specified, we will attempt to automatically detect the GCE project from "
    "metadata.")

tf.flags.DEFINE_string("master", None, "[Optional] TensorFlow master URL.")

flags.DEFINE_integer(
    "num_tpu_cores", 8,
    "Only used if `use_tpu` is True. Total number of TPU cores to use.")

flags.DEFINE_bool(
    "verbose_logging", False,
    "If true, all of the warnings related to data processing will be printed. "
    "A number of warnings are expected for a normal SQuAD evaluation.")

flags.DEFINE_bool(
    "version_2_with_negative", False,
    "If true, the SQuAD examples contain some that do not have an answer.")

flags.DEFINE_float(
    "null_score_diff_threshold", 0.0,
    "If null_score - best_non_null is greater than the threshold predict null.")

flags.DEFINE_bool(
    "do_retriever", False,
    "If True, use retriever to help reader to filte good doc - add by willy.")

flags.DEFINE_string(
    "retriever_model", None,
    "retriever model path - add by willy.")

flags.DEFINE_float(
    "retriever_weight", 0.0,
    "retriever weight - add by willy.")


flags.DEFINE_integer("retriever_ranker", 1,"Rank with retriever.")

flags.DEFINE_string("document_type","SQuAD", "There are three document types: (1)paragraphs in SQuAD (2)SQlite (DataBase) (3) Text - add by willy." )

flags.DEFINE_string("question_type","SQuAD", "There are three question types: (1) SQuAD (2)one_question (3) interactive." )

flags.DEFINE_string("question", None, "give question to predict - Willy Test.")

flags.DEFINE_string("db_file", None, "give path with data base file to set SQlite State - Willy Test.")

flags.DEFINE_string("question_table", None, "set table path - Willy Test.")

flags.DEFINE_string("excel_name", None ,"set excel name -Willy Test.")


class DecimalEncoder(json.JSONEncoder):
  def default(self, obj):
    if isinstance(obj, decimal.Decimal):
      return float(obj)
    return super(DecimalEncoder, self).default(obj)


class SquadExample(object):
  """A single training/test example for simple sequence classification.
     For examples without an answer, the start and end position are -1.
  """

  def __init__(self,
               qas_id,
               question_text,
               doc_id,                                                   #willy add
               doc_tokens,
               orig_answer_text=None,
               start_position=None,
               end_position=None,
               is_impossible=False):
    self.qas_id = qas_id
    self.question_text = question_text
    self.doc_id = doc_id                                                 #willy add
    self.doc_tokens = doc_tokens
    self.orig_answer_text = orig_answer_text
    self.start_position = start_position
    self.end_position = end_position
    self.is_impossible = is_impossible

  def __str__(self):
    return self.__repr__()

  def __repr__(self):
    s = ""
    s += "qas_id: %s" % (tokenization.printable_text(self.qas_id))
    s += ", question_text: %s" % (
        tokenization.printable_text(self.question_text))
    s += ", doc_id:[%s]" % (tokenization.printable_text(self.doc_id))     #willy add
    s += ", doc_tokens: [%s]" % (" ".join(self.doc_tokens))
    if self.start_position:
      s += ", start_position: %d" % (self.start_position)
    if self.start_position:
      s += ", end_position: %d" % (self.end_position)
    if self.start_position:
      s += ", is_impossible: %r" % (self.is_impossible)
    return s


class InputFeatures(object):
  """A single set of features of data."""

  def __init__(self,
               unique_id,
               example_index,
               doc_span_index,
               tokens,
               token_to_orig_map,
               token_is_max_context,
               input_ids,
               input_mask,
               segment_ids,
               start_position=None,
               end_position=None,
               is_impossible=None):
    self.unique_id = unique_id
    self.example_index = example_index
    self.doc_span_index = doc_span_index
    self.tokens = tokens
    self.token_to_orig_map = token_to_orig_map
    self.token_is_max_context = token_is_max_context
    self.input_ids = input_ids
    self.input_mask = input_mask
    self.segment_ids = segment_ids
    self.start_position = start_position
    self.end_position = end_position
    self.is_impossible = is_impossible

def TakeThird(val):
    return val[2]
    
    
def set_squad_examples(input_file,question):

    """Read a SQuAD json file into a list of SquadExample."""
    with tf.gfile.Open(input_file, "r") as reader:
        input_data = json.load(reader)["data"]
    def is_whitespace(c):
        if c == " " or c == "\t" or c == "\r" or c == "\n" or ord(c) == 0x202F:
            return True
        return False
        
    examples = []
    file = open("Output1.txt", "r")
    document = file.read()
    file.close()
    paragraphs = document.split('\n')
    paragraphs = list(filter(None, paragraphs))
    #-----------------------------------------------
    doc_tokensList = []
    for i , paragraph_text in enumerate(paragraphs):
        # paragraph_text = paragraph["context"]
        doc_tokens = []
        char_to_word_offset = []
        prev_is_whitespace = True
        for c in paragraph_text:
            if is_whitespace(c):
                prev_is_whitespace = True
            else:
                if prev_is_whitespace:
                    doc_tokens.append(c)
                else:
                    doc_tokens[-1] += c
                prev_is_whitespace = False
            char_to_word_offset.append(len(doc_tokens) - 1)
        doc_tokensList.append(doc_tokens)
    #-----------------------------------------------
    start_position = -1
    end_position = -1
    orig_answer_text = ""
    is_impossible = False
    for i, doc_tokens in enumerate(doc_tokensList):
        example = SquadExample(
            qas_id=str(uuid.uuid1()),
            question_text=question,
            doc_id=DOC2IDX[i],
            doc_tokens=doc_tokens,
            orig_answer_text=orig_answer_text,
            start_position=start_position,
            end_position=end_position,
            is_impossible=is_impossible)
        examples.append(example)
    
    '''    
    for entry in input_data:
        for paragraph in entry["paragraphs"]:
            for qa in paragraph["qas"]:
                #qas_id = qa["id"]
                # uuid reset by willy in 20190313
                qas_id = str(uuid.uuid1())
                question_text = qa["question"]
                start_position = -1
                end_position = -1
                orig_answer_text = ""
                is_impossible = False
                for doc_tokens in doc_tokensList:
                    example = SquadExample(
                        qas_id=qas_id,
                        question_text=question_text,
                        doc_tokens=doc_tokens,
                        orig_answer_text=orig_answer_text,
                        start_position=start_position,
                        end_position=end_position,
                        is_impossible=is_impossible)
                    print(example)
                    examples.append(example)
    '''         
    #-----------------------------------------------
    return examples        
    
def read_squad_examples(input_file, is_training):
  """Read a SQuAD json file into a list of SquadExample."""
  with tf.gfile.Open(input_file, "r") as reader:
    input_data = json.load(reader)["data"]

  def is_whitespace(c):
    if c == " " or c == "\t" or c == "\r" or c == "\n" or ord(c) == 0x202F:
      return True
    return False

  examples = []
  for entry in input_data:
    for paragraph in entry["paragraphs"]:
      paragraph_text = paragraph["context"]
      doc_tokens = []
      char_to_word_offset = []
      prev_is_whitespace = True
      for c in paragraph_text:
        if is_whitespace(c):
          prev_is_whitespace = True
        else:
          if prev_is_whitespace:
            doc_tokens.append(c)
          else:
            doc_tokens[-1] += c
          prev_is_whitespace = False
        char_to_word_offset.append(len(doc_tokens) - 1)

      for qa in paragraph["qas"]:
        qas_id = qa["id"]
        question_text = qa["question"]
        start_position = None
        end_position = None
        orig_answer_text = None
        is_impossible = False
        if is_training:

          if FLAGS.version_2_with_negative:
            is_impossible = qa["is_impossible"]
          if (len(qa["answers"]) != 1) and (not is_impossible):
            raise ValueError(
                "For training, each question should have exactly 1 answer.")
          if not is_impossible:
            answer = qa["answers"][0]
            orig_answer_text = answer["text"]
            answer_offset = answer["answer_start"]
            answer_length = len(orig_answer_text)
            start_position = char_to_word_offset[answer_offset]
            end_position = char_to_word_offset[answer_offset + answer_length -
                                               1]
            # Only add answers where the text can be exactly recovered from the
            # document. If this CAN'T happen it's likely due to weird Unicode
            # stuff so we will just skip the example.
            #
            # Note that this means for training mode, every example is NOT
            # guaranteed to be preserved.
            actual_text = " ".join(
                doc_tokens[start_position:(end_position + 1)])
            cleaned_answer_text = " ".join(
                tokenization.whitespace_tokenize(orig_answer_text))
            if actual_text.find(cleaned_answer_text) == -1:
              tf.logging.warning("Could not find answer: '%s' vs. '%s'",
                                 actual_text, cleaned_answer_text)
              continue
          else:
            start_position = -1
            end_position = -1
            orig_answer_text = ""

        example = SquadExample(
            qas_id=qas_id,
            question_text=question_text,
            doc_tokens=doc_tokens,
            orig_answer_text=orig_answer_text,
            start_position=start_position,
            end_position=end_position,
            is_impossible=is_impossible)
        examples.append(example)

  return examples

def convert_examples_to_features(examples, tokenizer, max_seq_length,
                                 doc_stride, max_query_length, is_training,
                                 output_fn):
  """Loads a data file into a list of `InputBatch`s."""

  unique_id = 1000000000

  for (example_index, example) in enumerate(examples):
    query_tokens = tokenizer.tokenize(example.question_text)


    if len(query_tokens) > max_query_length:
      query_tokens = query_tokens[0:max_query_length]

    tok_to_orig_index = []
    orig_to_tok_index = []
    all_doc_tokens = []
    for (i, token) in enumerate(example.doc_tokens):
      orig_to_tok_index.append(len(all_doc_tokens))
      sub_tokens = tokenizer.tokenize(token)
      for sub_token in sub_tokens:
        tok_to_orig_index.append(i)
        all_doc_tokens.append(sub_token)

    tok_start_position = None
    tok_end_position = None
    if is_training and example.is_impossible:
      tok_start_position = -1
      tok_end_position = -1
    if is_training and not example.is_impossible:
      tok_start_position = orig_to_tok_index[example.start_position]
      if example.end_position < len(example.doc_tokens) - 1:
        tok_end_position = orig_to_tok_index[example.end_position + 1] - 1
      else:
        tok_end_position = len(all_doc_tokens) - 1
      (tok_start_position, tok_end_position) = _improve_answer_span(
          all_doc_tokens, tok_start_position, tok_end_position, tokenizer,
          example.orig_answer_text)

    # The -3 accounts for [CLS], [SEP] and [SEP]
    max_tokens_for_doc = max_seq_length - len(query_tokens) - 3

    # We can have documents that are longer than the maximum sequence length.
    # To deal with this we do a sliding window approach, where we take chunks
    # of the up to our max length with a stride of `doc_stride`.
    _DocSpan = collections.namedtuple(  # pylint: disable=invalid-name
        "DocSpan", ["start", "length"])
    doc_spans = []
    start_offset = 0
    while start_offset < len(all_doc_tokens):
      length = len(all_doc_tokens) - start_offset
      if length > max_tokens_for_doc:
        length = max_tokens_for_doc
      doc_spans.append(_DocSpan(start=start_offset, length=length))
      if start_offset + length == len(all_doc_tokens):
        break
      start_offset += min(length, doc_stride)

    for (doc_span_index, doc_span) in enumerate(doc_spans):
      tokens = []
      token_to_orig_map = {}
      token_is_max_context = {}
      segment_ids = []
      tokens.append("[CLS]")
      segment_ids.append(0)
      for token in query_tokens:
        tokens.append(token)
        segment_ids.append(0)
      tokens.append("[SEP]")
      segment_ids.append(0)

      for i in range(doc_span.length):
        split_token_index = doc_span.start + i
        token_to_orig_map[len(tokens)] = tok_to_orig_index[split_token_index]

        is_max_context = _check_is_max_context(doc_spans, doc_span_index,
                                               split_token_index)
        token_is_max_context[len(tokens)] = is_max_context
        tokens.append(all_doc_tokens[split_token_index])
        segment_ids.append(1)
      tokens.append("[SEP]")
      segment_ids.append(1)

      input_ids = tokenizer.convert_tokens_to_ids(tokens)

      # The mask has 1 for real tokens and 0 for padding tokens. Only real
      # tokens are attended to.
      input_mask = [1] * len(input_ids)

      # Zero-pad up to the sequence length.
      while len(input_ids) < max_seq_length:
        input_ids.append(0)
        input_mask.append(0)
        segment_ids.append(0)

      assert len(input_ids) == max_seq_length
      assert len(input_mask) == max_seq_length
      assert len(segment_ids) == max_seq_length

      start_position = None
      end_position = None
      if is_training and not example.is_impossible:
        # For training, if our document chunk does not contain an annotation
        # we throw it out, since there is nothing to predict.
        doc_start = doc_span.start
        doc_end = doc_span.start + doc_span.length - 1
        out_of_span = False
        if not (tok_start_position >= doc_start and
                tok_end_position <= doc_end):
          out_of_span = True
        if out_of_span:
          start_position = 0
          end_position = 0
        else:
          doc_offset = len(query_tokens) + 2
          start_position = tok_start_position - doc_start + doc_offset
          end_position = tok_end_position - doc_start + doc_offset

      if is_training and example.is_impossible:
        start_position = 0
        end_position = 0

      if example_index < 10:
        tf.logging.info("*** Example ***")
        tf.logging.info("unique_id: %s" % (unique_id))
        tf.logging.info("example_index: %s" % (example_index))
        tf.logging.info("doc_span_index: %s" % (doc_span_index))
        tf.logging.info("tokens: %s" % " ".join(
            [tokenization.printable_text(x) for x in tokens]))
        tf.logging.info("token_to_orig_map: %s" % " ".join(
            ["%d:%d" % (x, y) for (x, y) in six.iteritems(token_to_orig_map)]))
        tf.logging.info("token_is_max_context: %s" % " ".join([
            "%d:%s" % (x, y) for (x, y) in six.iteritems(token_is_max_context)
        ]))
        tf.logging.info("input_ids: %s" % " ".join([str(x) for x in input_ids]))
        tf.logging.info(
            "input_mask: %s" % " ".join([str(x) for x in input_mask]))
        tf.logging.info(
            "segment_ids: %s" % " ".join([str(x) for x in segment_ids]))
        if is_training and example.is_impossible:
          tf.logging.info("impossible example")
        if is_training and not example.is_impossible:
          answer_text = " ".join(tokens[start_position:(end_position + 1)])
          tf.logging.info("start_position: %d" % (start_position))
          tf.logging.info("end_position: %d" % (end_position))
          tf.logging.info(
              "answer: %s" % (tokenization.printable_text(answer_text)))

      feature = InputFeatures(
          unique_id=unique_id,
          example_index=example_index,
          doc_span_index=doc_span_index,
          tokens=tokens,
          token_to_orig_map=token_to_orig_map,
          token_is_max_context=token_is_max_context,
          input_ids=input_ids,
          input_mask=input_mask,
          segment_ids=segment_ids,
          start_position=start_position,
          end_position=end_position,
          is_impossible=example.is_impossible)

      # Run callback
      output_fn(feature)

      unique_id += 1


def _improve_answer_span(doc_tokens, input_start, input_end, tokenizer,
                         orig_answer_text):
  """Returns tokenized answer spans that better match the annotated answer."""

  # The SQuAD annotations are character based. We first project them to
  # whitespace-tokenized words. But then after WordPiece tokenization, we can
  # often find a "better match". For example:
  #
  #   Question: What year was John Smith born?
  #   Context: The leader was John Smith (1895-1943).
  #   Answer: 1895
  #
  # The original whitespace-tokenized answer will be "(1895-1943).". However
  # after tokenization, our tokens will be "( 1895 - 1943 ) .". So we can match
  # the exact answer, 1895.
  #
  # However, this is not always possible. Consider the following:
  #
  #   Question: What country is the top exporter of electornics?
  #   Context: The Japanese electronics industry is the lagest in the world.
  #   Answer: Japan
  #
  # In this case, the annotator chose "Japan" as a character sub-span of
  # the word "Japanese". Since our WordPiece tokenizer does not split
  # "Japanese", we just use "Japanese" as the annotation. This is fairly rare
  # in SQuAD, but does happen.
  tok_answer_text = " ".join(tokenizer.tokenize(orig_answer_text))

  for new_start in range(input_start, input_end + 1):
    for new_end in range(input_end, new_start - 1, -1):
      text_span = " ".join(doc_tokens[new_start:(new_end + 1)])
      if text_span == tok_answer_text:
        return (new_start, new_end)

  return (input_start, input_end)


def _check_is_max_context(doc_spans, cur_span_index, position):
  """Check if this is the 'max context' doc span for the token."""

  # Because of the sliding window approach taken to scoring documents, a single
  # token can appear in multiple documents. E.g.
  #  Doc: the man went to the store and bought a gallon of milk
  #  Span A: the man went to the
  #  Span B: to the store and bought
  #  Span C: and bought a gallon of
  #  ...
  #
  # Now the word 'bought' will have two scores from spans B and C. We only
  # want to consider the score with "maximum context", which we define as
  # the *minimum* of its left and right context (the *sum* of left and
  # right context will always be the same, of course).
  #
  # In the example the maximum context for 'bought' would be span C since
  # it has 1 left context and 3 right context, while span B has 4 left context
  # and 0 right context.
  best_score = None
  best_span_index = None
  for (span_index, doc_span) in enumerate(doc_spans):
    end = doc_span.start + doc_span.length - 1
    if position < doc_span.start:
      continue
    if position > end:
      continue
    num_left_context = position - doc_span.start
    num_right_context = end - position
    score = min(num_left_context, num_right_context) + 0.01 * doc_span.length
    if best_score is None or score > best_score:
      best_score = score
      best_span_index = span_index

  return cur_span_index == best_span_index


def create_model(bert_config, is_training, input_ids, input_mask, segment_ids,
                 use_one_hot_embeddings):
  """Creates a classification model."""
  model = modeling.BertModel(
      config=bert_config,
      is_training=is_training,
      input_ids=input_ids,
      input_mask=input_mask,
      token_type_ids=segment_ids,
      use_one_hot_embeddings=use_one_hot_embeddings)

  final_hidden = model.get_sequence_output()

  final_hidden_shape = modeling.get_shape_list(final_hidden, expected_rank=3)
  batch_size = final_hidden_shape[0]
  seq_length = final_hidden_shape[1]
  hidden_size = final_hidden_shape[2]

  output_weights = tf.get_variable(
      "cls/squad/output_weights", [2, hidden_size],
      initializer=tf.truncated_normal_initializer(stddev=0.02))

  output_bias = tf.get_variable(
      "cls/squad/output_bias", [2], initializer=tf.zeros_initializer())

  final_hidden_matrix = tf.reshape(final_hidden,
                                   [batch_size * seq_length, hidden_size])
  logits = tf.matmul(final_hidden_matrix, output_weights, transpose_b=True)
  logits = tf.nn.bias_add(logits, output_bias)

  logits = tf.reshape(logits, [batch_size, seq_length, 2])
  logits = tf.transpose(logits, [2, 0, 1])

  unstacked_logits = tf.unstack(logits, axis=0)

  (start_logits, end_logits) = (unstacked_logits[0], unstacked_logits[1])

  return (start_logits, end_logits)


def model_fn_builder(bert_config, init_checkpoint, learning_rate,
                     num_train_steps, num_warmup_steps, use_tpu,
                     use_one_hot_embeddings):
  """Returns `model_fn` closure for TPUEstimator."""

  def model_fn(features, labels, mode, params):  # pylint: disable=unused-argument
    """The `model_fn` for TPUEstimator."""

    tf.logging.info("*** Features ***")
    for name in sorted(features.keys()):
      tf.logging.info("  name = %s, shape = %s" % (name, features[name].shape))

    unique_ids = features["unique_ids"]
    input_ids = features["input_ids"]
    input_mask = features["input_mask"]
    segment_ids = features["segment_ids"]

    is_training = (mode == tf.estimator.ModeKeys.TRAIN)

    (start_logits, end_logits) = create_model(
        bert_config=bert_config,
        is_training=is_training,
        input_ids=input_ids,
        input_mask=input_mask,
        segment_ids=segment_ids,
        use_one_hot_embeddings=use_one_hot_embeddings)

    tvars = tf.trainable_variables()

    initialized_variable_names = {}
    scaffold_fn = None
    if init_checkpoint:
      (assignment_map, initialized_variable_names
      ) = modeling.get_assignment_map_from_checkpoint(tvars, init_checkpoint)
      if use_tpu:

        def tpu_scaffold():
          tf.train.init_from_checkpoint(init_checkpoint, assignment_map)
          return tf.train.Scaffold()

        scaffold_fn = tpu_scaffold
      else:
        tf.train.init_from_checkpoint(init_checkpoint, assignment_map)

    tf.logging.info("**** Trainable Variables ****")
    for var in tvars:
      init_string = ""
      if var.name in initialized_variable_names:
        init_string = ", *INIT_FROM_CKPT*"
      tf.logging.info("  name = %s, shape = %s%s", var.name, var.shape,
                      init_string)

    output_spec = None
    if mode == tf.estimator.ModeKeys.TRAIN:
      seq_length = modeling.get_shape_list(input_ids)[1]

      def compute_loss(logits, positions):
        one_hot_positions = tf.one_hot(
            positions, depth=seq_length, dtype=tf.float32)
        log_probs = tf.nn.log_softmax(logits, axis=-1)
        loss = -tf.reduce_mean(
            tf.reduce_sum(one_hot_positions * log_probs, axis=-1))
        return loss

      start_positions = features["start_positions"]
      end_positions = features["end_positions"]

      start_loss = compute_loss(start_logits, start_positions)
      end_loss = compute_loss(end_logits, end_positions)

      total_loss = (start_loss + end_loss) / 2.0

      train_op = optimization.create_optimizer(
          total_loss, learning_rate, num_train_steps, num_warmup_steps, use_tpu)

      output_spec = tf.contrib.tpu.TPUEstimatorSpec(
          mode=mode,
          loss=total_loss,
          train_op=train_op,
          scaffold_fn=scaffold_fn)
    elif mode == tf.estimator.ModeKeys.PREDICT:
      predictions = {
          "unique_ids": unique_ids,
          "start_logits": start_logits,
          "end_logits": end_logits,
      }
      output_spec = tf.contrib.tpu.TPUEstimatorSpec(
          mode=mode, predictions=predictions, scaffold_fn=scaffold_fn)
    else:
      raise ValueError(
          "Only TRAIN and PREDICT modes are supported: %s" % (mode))

    return output_spec

  return model_fn


def input_fn_builder(input_file, seq_length, is_training, drop_remainder):
  """Creates an `input_fn` closure to be passed to TPUEstimator."""

  name_to_features = {
      "unique_ids": tf.FixedLenFeature([], tf.int64),
      "input_ids": tf.FixedLenFeature([seq_length], tf.int64),
      "input_mask": tf.FixedLenFeature([seq_length], tf.int64),
      "segment_ids": tf.FixedLenFeature([seq_length], tf.int64),
  }

  if is_training:
    name_to_features["start_positions"] = tf.FixedLenFeature([], tf.int64)
    name_to_features["end_positions"] = tf.FixedLenFeature([], tf.int64)

  def _decode_record(record, name_to_features):
    """Decodes a record to a TensorFlow example."""
    example = tf.parse_single_example(record, name_to_features)

    # tf.Example only supports tf.int64, but the TPU only supports tf.int32.
    # So cast all int64 to int32.
    for name in list(example.keys()):
      t = example[name]
      if t.dtype == tf.int64:
        t = tf.to_int32(t)
      example[name] = t

    return example

  def input_fn(params):
    """The actual input function."""
    batch_size = params["batch_size"]

    # For training, we want a lot of parallel reading and shuffling.
    # For eval, we want no shuffling and parallel reading doesn't matter.
    d = tf.data.TFRecordDataset(input_file)
    if is_training:
      d = d.repeat()
      d = d.shuffle(buffer_size=100)

    d = d.apply(
        tf.data.experimental.map_and_batch(
            lambda record: _decode_record(record, name_to_features),
            batch_size=batch_size,
            drop_remainder=drop_remainder))

    return d

  return input_fn


RawResult = collections.namedtuple("RawResult",
                                   ["unique_id", "start_logits", "end_logits"])


def write_predictions(all_examples, all_features, all_results, n_best_size,
                      max_answer_length, do_lower_case, output_prediction_file,
                      output_nbest_file, output_null_log_odds_file,
                      output_Aten_predict_file
                     ):
  """Write final predictions to the json file and log-odds of null if needed."""
  tf.logging.info("Writing predictions to: %s" % (output_prediction_file))
  tf.logging.info("Writing nbest to: %s" % (output_nbest_file))
  tf.logging.info("Writing nbest to: %s" % (output_Aten_predict_file))  

  example_index_to_features = collections.defaultdict(list)
  for feature in all_features:
    example_index_to_features[feature.example_index].append(feature)

  unique_id_to_result = {}
  tf.logging.info("length of all_results: %d" % (len(all_results)))
  for result in all_results:
    unique_id_to_result[result.unique_id] = result

  _PrelimPrediction = collections.namedtuple(  # pylint: disable=invalid-name
      "PrelimPrediction",
      ["feature_index", "start_index", "end_index", "start_logit", "end_logit"])


  # Willy Addd collections -> for results
  #-------------------------------------------------------------------------------  
  _AllPredictions = collections.namedtuple(  # pylint: disable=invalid-name
      "AllPredictions",
      ["question", "PredictListOneQues"])      

  _AllPredictResultsInOneQuestion = collections.namedtuple(  # pylint: disable=invalid-name
      "AllPredictResultsInOneQuestion",
      ["doc_text", "doc_id", "doc_score", "PredictListOneDoc"])

  _AllPredictResultsInOneDocument = collections.namedtuple(  # pylint: disable=invalid-name
      "AllPredictResultsInOneDocument",
      ["answer", "prob", "start", "end"])    

    
    
  _FinalResult = collections.namedtuple(  # pylint: disable=invalid-name
      "FinalResult",
      ["question", "text", "text_id", "ans", "prob"])
  _FinalResult2 = collections.namedtuple(  # pylint: disable=invalid-name
      "FinalResult2",
      ["question", "text", "ans", "prob"])

  _TempAllpredict_Layer1 = collections.namedtuple(  # pylint: disable=invalid-name 
      "TempAllpredict_Layer1",
      ["question" , "TempAllpredictList_Layer2"]) 

  _TempAllpredict_Layer2 = collections.namedtuple(  # pylint: disable=invalid-name 
      "TempAllpredict_Layer2",
      ["doc_id","doc_text","best_ans","best_prob"])
  #-------------------------------------------------------------------------------

  all_predictions = collections.OrderedDict()
  all_nbest_json = collections.OrderedDict()
  scores_diff_json = collections.OrderedDict()
  
 
  all_predicts = []
  all_predictsInOneQues = []
  quesList = []
  Aten_result_list = []
  Aten_result2_list = []
  TempAllpredictLayer1_list = []
  TempAllpredictLayer2_list = []
  best_answer=""
  best_prob=0.0
  ans_is_null = True
  
  ranker = retriever.get_class('tfidf')(tfidf_path=FLAGS.retriever_model)       
  for (example_index, example) in enumerate(all_examples):
    features = example_index_to_features[example_index]    
    
    if example_in_write_predictions == 1:
        print ("example idx:%d" %example_index)
        print("question in example from predict")
        print(example.question_text)
        print("doc_tokens in example from predict")
        print(example.doc_tokens)
        print('-'*60)
        print('\n')  
        

    doc_names, doc_scores = ranker.closest_docs( example.question_text, 10 )  
        
    prelim_predictions = []    
    
    # keep track of the minimum score of null start+end of position 0
    score_null = 1000000  # large and positive
    min_null_feature_index = 0  # the paragraph slice with min mull score
    null_start_logit = 0  # the start logit at the slice with min null score
    null_end_logit = 0  # the end logit at the slice with min null score
    for (feature_index, feature) in enumerate(features):
        
      result = unique_id_to_result[feature.unique_id]
      start_indexes = _get_best_indexes(result.start_logits, n_best_size)
      end_indexes = _get_best_indexes(result.end_logits, n_best_size)
        
      # if we could have irrelevant answers, get the min score of irrelevant
      if FLAGS.version_2_with_negative:
        feature_null_score = result.start_logits[0] + result.end_logits[0]
        if feature_null_score < score_null:
          score_null = feature_null_score
          min_null_feature_index = feature_index
          null_start_logit = result.start_logits[0]
          null_end_logit = result.end_logits[0]
      for start_index in start_indexes:
        for end_index in end_indexes:
          # We could hypothetically create invalid predictions, e.g., predict
          # that the start of the span is in the question. We throw out all
          # invalid predictions.
          if start_index >= len(feature.tokens):
            continue
          if end_index >= len(feature.tokens):
            continue
          if start_index not in feature.token_to_orig_map:
            continue
          if end_index not in feature.token_to_orig_map:
            continue
          if not feature.token_is_max_context.get(start_index, False):
            continue
          if end_index < start_index:
            continue
          length = end_index - start_index + 1
          if length > max_answer_length:
            continue
          prelim_predictions.append(
              _PrelimPrediction(
                  feature_index=feature_index,
                  start_index=start_index,
                  end_index=end_index,
                  start_logit=result.start_logits[start_index],
                  end_logit=result.end_logits[end_index]))
        
        
    if FLAGS.version_2_with_negative:
      prelim_predictions.append(
          _PrelimPrediction(
              feature_index=min_null_feature_index,
              start_index=0,
              end_index=0,
              start_logit=null_start_logit,
              end_logit=null_end_logit))   
    
    
    prelim_predictions = sorted(
        prelim_predictions,
        key=lambda x: (x.start_logit + x.end_logit),
        reverse=True)

    _NbestPrediction = collections.namedtuple(  # pylint: disable=invalid-name
        "NbestPrediction", ["text", "start_logit", "end_logit"])
    

    seen_predictions = {}
    nbest = []
    
    for pred in prelim_predictions:
      if len(nbest) >= n_best_size:
        break
      feature = features[pred.feature_index]
      if pred.start_index > 0:  # this is a non-null prediction
        tok_tokens = feature.tokens[pred.start_index:(pred.end_index + 1)]
        orig_doc_start = feature.token_to_orig_map[pred.start_index]
        orig_doc_end = feature.token_to_orig_map[pred.end_index]
        orig_tokens = example.doc_tokens[orig_doc_start:(orig_doc_end + 1)]
        tok_text = " ".join(tok_tokens)

        # De-tokenize WordPieces that have been split off.
        tok_text = tok_text.replace(" ##", "")
        tok_text = tok_text.replace("##", "")

        # Clean whitespace
        tok_text = tok_text.strip()
        tok_text = " ".join(tok_text.split())
        orig_text = " ".join(orig_tokens)

        final_text = get_final_text(tok_text, orig_text, do_lower_case)
        
        if final_text in seen_predictions:
          continue            
        seen_predictions[final_text] = True    
        
      else:
        final_text = ""
        seen_predictions[final_text] = True

      nbest.append(
          _NbestPrediction(
              text=final_text,
              start_logit=pred.start_logit,
              end_logit=pred.end_logit))            


    # if we didn't inlude the empty option in the n-best, inlcude it
    if FLAGS.version_2_with_negative:
      if "" not in seen_predictions:
        nbest.append(
            _NbestPrediction(
                text="", start_logit=null_start_logit,
                end_logit=null_end_logit))
        
    # In very rare edge cases we could have no valid predictions. So we
    # just create a nonce prediction in this case to avoid failure.
    if not nbest:
      nbest.append(
          _NbestPrediction(text="empty", start_logit=0.0, end_logit=0.0))

    assert len(nbest) >= 1
     
    total_scores = []
    best_non_null_entry = None
    for entry in nbest:
      total_scores.append(entry.start_logit + entry.end_logit)
      if not best_non_null_entry:
        if entry.text:
          best_non_null_entry = entry

    #參考
    probs = _compute_softmax(total_scores)
    
    nbest_json = []
    for i, entry in enumerate(nbest):
      output = collections.OrderedDict()
      output["text"] = entry.text
      output["probability"] = probs[i]
      output["start_logit"] = entry.start_logit
      output["end_logit"] = entry.end_logit
      nbest_json.append(output)

    # doc_names, doc_scores
    #----------------------------------------------
    # presupposition : Question is in order
    #"question", "PredictResults"
    if example.question_text not in quesList :
        if len(quesList)!=0 :
            #1. Save to all predicts
            #print('all_predictsInOneQues-')
            #print(all_predictsInOneQues)
            temp = copy.deepcopy(all_predictsInOneQues)
            #print('temp')
            #print(temp)
            all_predicts.append(
                _AllPredictions(
                    question=quesList[-1], 
                    PredictListOneQues=temp
                )
            ) 
            #2.TODO : Find the result (move to outside)
            #3. reset all_predictsInOneQues
            all_predictsInOneQues.clear()
            
        #. Add to questList
        quesList.append(example.question_text)
    #----------------------------------------------     
    

    # save answer dataset
    #----------------------------------------------
    all_predictsInOneDoc = [] 
    #print('go to (1)')
    for i, entry in enumerate(nbest):        
        if i==2:
            #print('In state 2')
            break
        tp_answer = entry.text    
        if i==0 :
            if tp_answer.isspace() or not tp_answer:
                #print('In state 0,tp_ans: %s' %tp_answer)
                continue
        if i == 1 and len(all_predictsInOneDoc)!=0:
            #print('In state 1,tp_ans: %s' %tp_answer)
            break
        #print('In state set pridict. tp_ans: %s' %tp_answer )    
        all_predictsInOneDoc.append(
            _AllPredictResultsInOneDocument(
                answer=entry.text, 
                prob=Decimal(probs[i]),
                start = entry.start_logit,
                end = entry.end_logit
            )
        )
    
    #print('go to (2)')  
    #----------------------------------------------
    # End of save answer dataset
    if predict_result_index == 1:
        for i, entry in enumerate(all_predictsInOneDoc): 
            print('index:%d' %i)
            print("answer: %s" %(entry.answer))
            print("prob: %s" %(entry.prob))
            print("start: %s" %(entry.start))
            print("end: %s" %(entry.end))
            print('\n')
        print('-'*15)
        print('\n')
    # append predicts to OneQues
    #print('go to (3)')
    #----------------------------------------------
    tp_docscore = 0.0
    if example.doc_id in doc_names :
        tp_docindex = doc_names.index(example.doc_id)
        tp_docscore = doc_scores [tp_docindex]
        #print('go to (4)')
    
    #print('go to (5)')    
    #print('all_predictsInOneQues-in set')
    #print(all_predictsInOneQues)    
    all_predictsInOneQues.append(
        _AllPredictResultsInOneQuestion(
            doc_text=example.doc_tokens,
            doc_id=example.doc_id,
            doc_score=tp_docscore,
            PredictListOneDoc=all_predictsInOneDoc
        )
    )
    #print('go to (6)')
    
    #print('all_predictsInOneQues-in set')
    #print(all_predictsInOneQues)
    #----------------------------------------------    
    
    # if example is examples last data
    if example == all_examples[-1] :
        all_predicts.append(
            _AllPredictions(question=example.question_text,PredictListOneQues=all_predictsInOneQues))             
    #----------------------------------------------
      
        
    assert len(nbest_json) >= 1
    if not FLAGS.version_2_with_negative:
      all_predictions[example.qas_id] = nbest_json[0]["text"]
    else:
      # predict "" iff the null score - the score of best non-null > threshold
      if best_non_null_entry == None :
          score_diff = FLAGS.null_score_diff_threshold + 1.0
      else:
        score_diff = score_null - best_non_null_entry.start_logit - (
            best_non_null_entry.end_logit)
      scores_diff_json[example.qas_id] = score_diff
      if score_diff > FLAGS.null_score_diff_threshold:
        all_predictions[example.qas_id] = ""
      else:
        all_predictions[example.qas_id] = best_non_null_entry.text
        
        
    all_nbest_json[example.qas_id] = nbest_json
    

  
  #TODO: Find the best answer from Aten collections
  #----------------------------------------------    
  const_AtenQuest_index =  [3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,4,3,6,5,5,5,5,5,5,5,5,5,3,5,4,5,5,5,5,5,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1]   
  const_AtenIntent_index = [1,1,1,0,1,1,1,1,1,1,1,1,1,1,1,1,1,0,1,1,1,1,1,1,1,0,1,1,1,0,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1]  
  excel_NOtGoodAns_index = [0,0,0,3,0,0,0,0,0,0,0,0,0,0,0,0,0,3,0,0,0,0,4,2,0,5,3,0,2,5,0,0,2,0,3,1,0,2,0,4,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
                         
    
  excel_count = 0 
  excel_index = 1 
  excel_Answer_count = const_AtenQuest_index[excel_index-1]
  excel_Intent_count = const_AtenIntent_index[excel_index-1]
  excel_NOtGoodAns_count = excel_NOtGoodAns_index[excel_index-1]
  wb = Workbook()
  ws = wb.active    
  retriever_weight = FLAGS.retriever_weight    

  #print('len of all_predicts:%d' %len(all_predicts))
    
  for i, entry_predicts in enumerate(all_predicts):
    tp_ques = entry_predicts.question   
    QuesList = entry_predicts.PredictListOneQues     
    
    #print("ques: %s" %(tp_ques))
    QuesList.sort(key=TakeThird, reverse=True)
    
    #print('len with QuesList:%d' %len(QuesList))
    best_doc = QuesList[0].doc_text

    entry_OneDoc = QuesList [0].PredictListOneDoc
    
    for k, entry_OneAns in enumerate(entry_OneDoc):
        #print('index:%d' %k)
        best_prob = Decimal(entry_OneAns.prob)
        best_ans = entry_OneAns.answer
        #print('Ans_ans:%s' %(entry_OneAns.answer))
        #print('Ans_prob:%e , start:%e , end:%e' %(entry_OneAns.prob , entry_OneAns.start , entry_OneAns.end))  
        
        str_result=""
        for word in best_doc:
            str_result= str_result + " " + word

        Aten_result2_list.append(
           _FinalResult2(
                question = tp_ques,
                text     = str_result,
                ans      = best_ans,
                prob     = best_prob
            )
        )
        print('ques: %s' %tp_ques)
        print('text: %s' %str_result)
        print('ans: %s' %best_ans)
        print('prob: %s' %best_prob)
        
        
        if excel_Answer_count == excel_count+1 : 
            print('-'*15)
            print('\n') 
        else :
            print('-'*5)            
        
        if excel_Answer_count == excel_count : 
            ws['C' + str(excel_index)] = excel_Answer_count
            ws['D' + str(excel_index)] = excel_NOtGoodAns_count
            ws['F' + str(excel_index)] = excel_Intent_count
        
            excel_index = excel_index+1
            excel_Answer_count = const_AtenQuest_index[excel_index-1]    
            excel_NOtGoodAns_count = excel_NOtGoodAns_index[excel_index-1]
            excel_Intent_count = const_AtenIntent_index[excel_index-1]   
            excel_count = 0            

            
            
        
        if excel_index <= len(const_AtenQuest_index) :
            index_str = chr(73+excel_count) + str(excel_index) 
            ws[index_str] = best_prob
            excel_count  = excel_count + 1


    
    '''
    for j , entry_OneDoc in enumerate(QuesList):
        print("DocIndex= %d " %(j))
        
        print('DocID: %s , DocScore:%e' %(entry_OneDoc.doc_id, entry_OneDoc.doc_score))
            
        print('DocText:')
        print(entry_OneDoc.doc_text)
        
        for k, entry_OneAns in enumerate(entry_OneDoc):
            print('index:%d' %k)
            print(entry_OneAns)
            tp_now_prob = Decimal(entry_OneAns.prob)
            print('Ans_ans:%s' %(entry_OneAns.answer))
            print('Ans_prob:%e , start:%e , end:%e' %(entry_OneAns.prob , entry_OneAns.start , entry_OneAns.end))
        # do 1a2b ....
        # TODO:
        
        # tp_now_prob = Decimal(retriever_weight)*Decimal(doc_score) + Decimal(1.0-retriever_weight)*Decimal(tp_now_prob)
    '''    
        
    
    '''
    for j, entry_OneQues in enumerate(QuesList):
        tp_text = entry_OneQues.doc_text
        doc_id = entry_OneQues.doc_id
        doc_score = entry_OneQues.doc_score        
        
        DocList = entry_OneQues.PredictListOneDoc
    ''' 
    
  # 2.choice best answer in one document
  # destination: 
  # way :     
  #---------------------------------------#             
  '''
  for i, entry_predicts in enumerate(all_predicts):
    tp_ques = entry_predicts.question   
    QuesList = entry_predicts.PredictListOneQues
    
    if checkState_in_AtenResult2 == 1:
        print("Ques:")
        print("Ques_ID=%d, tp_ques=%s" %(i,tp_ques) )    
    
    if ranker!= None :        
        doc_names, doc_scores = ranker.closest_docs( tp_ques, len(QuesList) )  
        table = prettytable.PrettyTable(
            ['Rank', 'Doc Id', 'Doc Score']
        )        
        for i in range(len(doc_names)):
            table.add_row([i + 1, doc_names[i], '%.5g' % doc_scores[i]])
    
    tp_no_answer = entry_predicts.no_answer
    best_ans = ""
    best_prob = 0.0
    best_doc = ""
    best_Docidx = 0
    
    tp_no_answer_ori = entry_predicts.no_answer
    best_ans_ori = ""
    best_prob_ori = 0.0
    best_doc_ori = ""
    best_Docidx_ori = 0  
    
    if checkState_in_AtenResult2 == 1:
        print('len of QuesList:%d' %len(QuesList))
        
    for j, entry_OneQues in enumerate(QuesList):
        tp_text = entry_OneQues.doc_text
        doc_id = entry_OneQues.doc_id
        DocList = entry_OneQues.PredictListOneDoc
        # Doc score
        doc_score = 0.0
        if doc_id in doc_names:
            doc_score = doc_scores[doc_names.index(doc_id)]
        
        
        #check state
        #--------------------------------------------------------------------------
        if checkState_in_AtenResult2 == 1:
            print("DocIndex= %d " %(j))
            
            print('DocID: %s' %(entry_OneQues.doc_id))
            
            print('DocText:')
            print(entry_OneQues.doc_text)
        
        #--------------------------------------------------------------------------
        
        
        if checkState_in_AtenResult2 == 1:
            print('len of DocList:%d' %len(DocList))
        #
        #---------------------------------------#            
        for k, entry_Doc in enumerate(DocList):
            
            tp_now_prob = Decimal(entry_Doc.prob)
            tp_now_prob2 = Decimal(entry_Doc.prob)
          
            #print('retriever_weight:%f, tp_now_prob:%f, doc_score:%f' %(retriever_weight, tp_now_prob, doc_score))
            if FLAGS.do_retriever:
                tp_now_prob = Decimal(retriever_weight)*Decimal(doc_score) + Decimal(1.0-retriever_weight)*Decimal(tp_now_prob)
                
            if checkState_in_AtenResult2 == 1:
                print("Weight: Ans_id=%d, Answer=%s , prob=%e" %(k, entry_Doc.answer , tp_now_prob))
                print("Original: Ans_id=%d, Answer=%s , prob=%e" %(k, entry_Doc.answer , tp_now_prob2))
            #
            #-----------------------------------#
            if tp_no_answer == False and k == 0:
                #
                #-------------------------------#
                if checkState_in_AtenResult == 1:
                    print("In Doc State 1, Weight: Ans_id=%d, Answer=%s , prob=%e" %(k, entry_Doc.answer , tp_now_prob)) 
                    print("In Doc State 1, Original : Ans_id=%d, Answer=%s , prob=%e" %(k, entry_Doc.answer , tp_now_prob2))
                # do weight:
                if entry_Doc.answer.strip() != "" and entry_Doc.answer.strip() != " " and tp_now_prob > best_prob:
                    if checkState_in_AtenResult == 1:
                        print("Reset answer:")
                        print("Weight: original data: best_ans: %s, best_prob=%e,best_Docidx=%d" %(best_ans, best_prob,best_Docidx))
    
                    best_ans = entry_Doc.answer
                    best_prob = tp_now_prob
                    best_doc = tp_text
                    best_Docidx = j
                                        
                    if checkState_in_AtenResult == 1:
                        print("change data: best_ans: %s, best_prob=%e,best_Docidx=%d" %(best_ans, best_prob,best_Docidx))
                
                # do original:
                if entry_Doc.answer.strip() != "" and entry_Doc.answer.strip() != " " and tp_now_prob2 > best_prob_ori:
                    if checkState_in_AtenResult == 1:
                        print("Reset answer:")
                        print("Orginal: original data: best_ans: %s, best_prob=%e,best_Docidx=%d" %(best_ans_ori, best_prob_ori,best_Docidx_ori))                    
                    best_ans_ori = entry_Doc.answer
                    best_prob_ori = tp_now_prob2
                    best_doc_ori = tp_text
                    best_Docidx_ori = j
                    if checkState_in_AtenResult == 1:
                        print("change data: best_ans: %s, best_prob=%e,best_Docidx=%d" %(best_ans_ori, best_prob_ori,best_Docidx_ori))
                    
            #-------------------------------#            
            #
            #-----------------------------------#
            elif entry_Doc.answer.strip() != "" and entry_Doc.answer.strip() != " " and tp_no_answer == True and k == 1:
                #
                #-------------------------------#
                if checkState_in_AtenResult == 1:
                    print("In Doc State 2, Weight: Ans_id=%d, Answer=%s , prob=%e" %(k, entry_Doc.answer , tp_now_prob)) 
                    print("In Doc State 2, Original : Ans_id=%d, Answer=%s , prob=%e" %(k, entry_Doc.answer , tp_now_prob2))
           
                if tp_now_prob > best_prob:
                    if checkState_in_AtenResult == 1:
                        print("Reset answer:")
                        print("original data: best_ans: %s, best_prob=%e, best_Docidx=%d" %(best_ans_ori, best_prob, best_Docidx))
                    best_ans = entry_Doc.answer
                    best_prob = tp_now_prob
                    best_doc = tp_text
                    best_Docidx = j
                    if checkState_in_AtenResult == 1:
                        print("change data: best_ans: %s, best_prob=%e,best_Docidx=%d" %(best_ans, best_prob,best_Docidx))
                if tp_now_prob2 > best_prob_ori:
                    if checkState_in_AtenResult == 1:
                        print("Reset answer:")
                        print("Orginal: original data: best_ans: %s, best_prob=%e,best_Docidx=%d" %(best_ans_ori, best_prob_ori,best_Docidx_ori))                    
                    best_ans_ori = entry_Doc.answer
                    best_prob_ori = tp_now_prob2
                    best_doc_ori = tp_text
                    best_Docidx_ori = j          
                    if checkState_in_AtenResult == 1:
                        print("change data: best_ans: %s, best_prob=%e,best_Docidx=%d" %(best_ans_ori, best_prob_ori,best_Docidx_ori))
                    
                #-------------------------------#
            #-----------------------------------#
            else:
                if checkState_in_AtenResult==1:
                    print(" In Doc State 3, Weight: Ans_id=%d, Answer=%s , prob=%e" %(k, entry_Doc.answer , tp_now_prob))
                    print(" In Doc State 3, Orginal: Ans_id=%d, Answer=%s , prob=%e" %(k, entry_Doc.answer , tp_now_prob2))
        #-----------------------------------# end of for Doc_List
        TempAllpredictLayer2_list.append(
            _TempAllpredict_Layer2(
                doc_id = best_Docidx_ori ,
                doc_text = best_doc_ori ,
                best_ans = best_ans_ori ,
                best_prob = best_prob_ori
            )
        )
    TempAllpredictLayer1_list.append(
        _TempAllpredict_Layer1(
            question = tp_ques,
            TempAllpredictList_Layer2 = TempAllpredictLayer2_list
        )
    )              
    #---------------------------------------# end of for Ques_List
       
    
    str_result=""
    for word in best_doc:
        str_result= str_result + " " + word
    Aten_result2_list.append(
        _FinalResult2(
            question = tp_ques,
            text     = str_result,
            ans      = best_ans,
            prob     = best_prob
        )
    )
    if excel_Answer_count == excel_count : 
        ws['C' + str(excel_index)] = excel_Answer_count
        ws['D' + str(excel_index)] = excel_NOtGoodAns_count
        ws['F' + str(excel_index)] = excel_Intent_count
        excel_index = excel_index+1
        excel_Answer_count = const_AtenQuest_index[excel_index-1]    
        excel_NOtGoodAns_count = excel_NOtGoodAns_index[excel_index-1]
        excel_Intent_count = const_AtenIntent_index[excel_index-1]   
        excel_count = 0
        
    if excel_index <= len(const_AtenQuest_index) :
        index_str = chr(73+excel_count) + str(excel_index) 
        ws[index_str] = best_prob
        excel_count  = excel_count + 1
    '''      


    #-------------------------------------------------# 

  ws['A60'] = 'All'
  ws['A61'] = '40QA'
  
  ws['B59'] = 'Right answer'
  ws['B60'] = '=SUM(B1:B40)+SUM(A41:A58)'
  ws['B61'] = '=SUM(B1:B40)'

  ws['C59'] = 'All answer'
  ws['C60'] = '=SUM(C1:C58)-SUM(D1:D40)'
  ws['C61'] = '=SUM(C1:C40)-SUM(D1:D40)'
    
  ws['E59'] = 'Right Intent'
  ws['E60'] = '=SUM(E1:E40)+SUM(A41:A58)'
  ws['E61'] = '=SUM(E1:E40)'    

  ws['F59'] = 'All intent'
  ws['F60'] = '=SUM(F1:F40)+SUM(C41:C58)'
  ws['F61'] = '=SUM(F1:F40)'    
    
    
  ws['G59'] = 'answer prob'
  ws['G60'] = '=B60/C60'
  ws['G61'] = '=B61/C61'    

  ws['H59'] = 'Intent prob'
  ws['H60'] = '=E60/F60'
  ws['H61'] = '=E61/F61'        
    
  wb.save(FLAGS.excel_name + '.xlsx')
  print('\n') 
  

  
    
  with tf.gfile.GFile(output_Aten_predict_file, "w") as writer:
    writer.write(json.dumps(Aten_result2_list, indent=4,cls=DecimalEncoder) + "\n")

  with tf.gfile.GFile(output_prediction_file, "w") as writer:
    writer.write(json.dumps(all_predictions, indent=4) + "\n")

  with tf.gfile.GFile(output_nbest_file, "w") as writer:
    writer.write(json.dumps(all_nbest_json, indent=4) + "\n")

  if FLAGS.version_2_with_negative:
    with tf.gfile.GFile(output_null_log_odds_file, "w") as writer:
      writer.write(json.dumps(scores_diff_json, indent=4) + "\n")


def get_final_text(pred_text, orig_text, do_lower_case):
  """Project the tokenized prediction back to the original text."""

  # When we created the data, we kept track of the alignment between original
  # (whitespace tokenized) tokens and our WordPiece tokenized tokens. So
  # now `orig_text` contains the span of our original text corresponding to the
  # span that we predicted.
  #
  # However, `orig_text` may contain extra characters that we don't want in
  # our prediction.
  #
  # For example, let's say:
  #   pred_text = steve smith
  #   orig_text = Steve Smith's
  #
  # We don't want to return `orig_text` because it contains the extra "'s".
  #
  # We don't want to return `pred_text` because it's already been normalized
  # (the SQuAD eval script also does punctuation stripping/lower casing but
  # our tokenizer does additional normalization like stripping accent
  # characters).
  #
  # What we really want to return is "Steve Smith".
  #
  # Therefore, we have to apply a semi-complicated alignment heruistic between
  # `pred_text` and `orig_text` to get a character-to-charcter alignment. This
  # can fail in certain cases in which case we just return `orig_text`.

  def _strip_spaces(text):
    ns_chars = []
    ns_to_s_map = collections.OrderedDict()
    for (i, c) in enumerate(text):
      if c == " ":
        continue
      ns_to_s_map[len(ns_chars)] = i
      ns_chars.append(c)
    ns_text = "".join(ns_chars)
    return (ns_text, ns_to_s_map)

  # We first tokenize `orig_text`, strip whitespace from the result
  # and `pred_text`, and check if they are the same length. If they are
  # NOT the same length, the heuristic has failed. If they are the same
  # length, we assume the characters are one-to-one aligned.
  tokenizer = tokenization.BasicTokenizer(do_lower_case=do_lower_case)

  tok_text = " ".join(tokenizer.tokenize(orig_text))

  start_position = tok_text.find(pred_text)
  if start_position == -1:
    if FLAGS.verbose_logging:
      tf.logging.info(
          "Unable to find text: '%s' in '%s'" % (pred_text, orig_text))
    return orig_text
  end_position = start_position + len(pred_text) - 1

  (orig_ns_text, orig_ns_to_s_map) = _strip_spaces(orig_text)
  (tok_ns_text, tok_ns_to_s_map) = _strip_spaces(tok_text)

  if len(orig_ns_text) != len(tok_ns_text):
    if FLAGS.verbose_logging:
      tf.logging.info("Length not equal after stripping spaces: '%s' vs '%s'",
                      orig_ns_text, tok_ns_text)
    return orig_text

  # We then project the characters in `pred_text` back to `orig_text` using
  # the character-to-character alignment.
  tok_s_to_ns_map = {}
  for (i, tok_index) in six.iteritems(tok_ns_to_s_map):
    tok_s_to_ns_map[tok_index] = i

  orig_start_position = None
  if start_position in tok_s_to_ns_map:
    ns_start_position = tok_s_to_ns_map[start_position]
    if ns_start_position in orig_ns_to_s_map:
      orig_start_position = orig_ns_to_s_map[ns_start_position]

  if orig_start_position is None:
    if FLAGS.verbose_logging:
      tf.logging.info("Couldn't map start position")
    return orig_text

  orig_end_position = None
  if end_position in tok_s_to_ns_map:
    ns_end_position = tok_s_to_ns_map[end_position]
    if ns_end_position in orig_ns_to_s_map:
      orig_end_position = orig_ns_to_s_map[ns_end_position]

  if orig_end_position is None:
    if FLAGS.verbose_logging:
      tf.logging.info("Couldn't map end position")
    return orig_text

  output_text = orig_text[orig_start_position:(orig_end_position + 1)]
  return output_text


def _get_best_indexes(logits, n_best_size):
  """Get the n-best logits from a list."""
  index_and_score = sorted(enumerate(logits), key=lambda x: x[1], reverse=True)

  best_indexes = []
  for i in range(len(index_and_score)):
    if i >= n_best_size:
      break
    best_indexes.append(index_and_score[i][0])
  return best_indexes


def _compute_softmax(scores):
  """Compute softmax probability over raw logits."""
  if not scores:
    return []

  max_score = None
  for score in scores:
    if max_score is None or score > max_score:
      max_score = score

  exp_scores = []
  total_sum = 0.0
  for score in scores:
    x = math.exp(score - max_score)
    exp_scores.append(x)
    total_sum += x

  probs = []
  for score in exp_scores:
    probs.append(score / total_sum)
  return probs


class FeatureWriter(object):
  """Writes InputFeature to TF example file."""

  def __init__(self, filename, is_training):
    self.filename = filename
    self.is_training = is_training
    self.num_features = 0
    self._writer = tf.python_io.TFRecordWriter(filename)

  def process_feature(self, feature):
    """Write a InputFeature to the TFRecordWriter as a tf.train.Example."""
    self.num_features += 1

    def create_int_feature(values):
      feature = tf.train.Feature(
          int64_list=tf.train.Int64List(value=list(values)))
      return feature

    features = collections.OrderedDict()
    features["unique_ids"] = create_int_feature([feature.unique_id])
    features["input_ids"] = create_int_feature(feature.input_ids)
    features["input_mask"] = create_int_feature(feature.input_mask)
    features["segment_ids"] = create_int_feature(feature.segment_ids)

    if self.is_training:
      features["start_positions"] = create_int_feature([feature.start_position])
      features["end_positions"] = create_int_feature([feature.end_position])
      impossible = 0
      if feature.is_impossible:
        impossible = 1
      features["is_impossible"] = create_int_feature([impossible])

    tf_example = tf.train.Example(features=tf.train.Features(feature=features))
    #print("tf_example:")
    #print(tf_example)
    self._writer.write(tf_example.SerializeToString())

  def close(self):
    self._writer.close()


def validate_flags_or_throw(bert_config):
  """Validate the input FLAGS or throw an exception."""
  tokenization.validate_case_matches_checkpoint(FLAGS.do_lower_case,
                                                FLAGS.init_checkpoint)

  if not FLAGS.do_train and not FLAGS.do_predict:
    raise ValueError("At least one of `do_train` or `do_predict` must be True.")

  if FLAGS.do_train:
    if not FLAGS.train_file:
      raise ValueError(
          "If `do_train` is True, then `train_file` must be specified.")
  if FLAGS.do_predict:
    if not FLAGS.predict_file:
      raise ValueError(
          "If `do_predict` is True, then `predict_file` must be specified.")

  if FLAGS.max_seq_length > bert_config.max_position_embeddings:
    raise ValueError(
        "Cannot use sequence length %d because the BERT model "
        "was only trained up to sequence length %d" %
        (FLAGS.max_seq_length, bert_config.max_position_embeddings))

  if FLAGS.max_seq_length <= FLAGS.max_query_length + 3:
    raise ValueError(
        "The max_seq_length (%d) must be greater than max_query_length "
        "(%d) + 3" % (FLAGS.max_seq_length, FLAGS.max_query_length))
  
  # Retriever - added by Willy 
  if FLAGS.do_retriever:
    if not FLAGS.retriever_model:
        raise ValueError("You have to set retriever model(give the path) when you set do_retriever to Yes.")
    if FLAGS.document_type != 'Sqlite' or FLAGS.db_file == None :
        raise ValueError("You have to set document_type to Sqlit and set the db_file when you set do_retriever to Yes.")
  
  # TODO : think a mechanism to chek these key word
  '''
  if FLAGS.document_type is 'SQlite':
    # TODO: set database
  elif FLAGS.document_type is 'Text':
    # TODO: set text file
  elif FLAGS.document_type is 'SQuAD':
    # is original method
  else :
    raise ValueError(
        "You have to set correct document_type: (1)'SQlite' (2)'Text' (3)SQuAD.")
'''

def read_squad_documents(input_file):
    """Read a SQuAD json file into a list of SquadExample."""
    with tf.gfile.Open(input_file, "r") as reader:
        input_data = json.load(reader)["data"]  
    documents = []
    for entry in input_data:
        for paragraph in entry["paragraphs"]:
            documents.append(paragraph["context"])
            
    return documents


def read_sqlite_documents(input_file):
    # TODO
    
    with db_class(input_file) as doc_db:
        doc_ids = doc_db.get_doc_ids()
        for ids in doc_ids:
            documents.append(doc_db.get_doc_text(ids))
    DOC2IDX = {doc_id: i for i, doc_id in enumerate(doc_ids)}
    return DOC2IDX, documents


def read_text_documents(input_file):
    examples = []
    file = open(input_file, "r")
    documents = file.read()
    file.close()
    documents_split = documents.split('\n')
    documents_final = list(filter(None, documents))
    return documents_final

def read_squad_question(input_file):
    questions = []
    """Read a SQuAD json file into a list of SquadExample."""
    with tf.gfile.Open(input_file, "r") as reader:    
        input_data = json.load(reader)["data"]  
        for entry in input_data:
            for paragraph in entry["paragraphs"]:
                for qa in paragraph["qas"]:
                    questions.append(qa["question"])
    return questions

def set_eval_examples(questions, DOC2IDX):
    def is_whitespace(c):
        if c == " " or c == "\t" or c == "\r" or c == "\n" or ord(c) == 0x202F:
            return True
        return False

    eval_examples = []
    temp_list = []
    print ('Show 2 DOC2IDX, len=%d'%(len(DOC2IDX)))
    print (DOC2IDX)
    for i, DOCID in enumerate(DOC2IDX) :
        print('ID:%d ,doc:%s' %(i,DOCID))
        temp_list.append(DOCID)
    
    for question in questions:
    #-------------------------questions - Start---------------------------#        
        question_text = question
        start_position = -1
        end_position = -1
        orig_answer_text = ""
        is_impossible = False

        #-------------documents - Start--------------#
        for i , paragraph_text in enumerate(documents):
            paragraph_text = paragraph_text
        #-------paragraphs - Start-------#
            doc_tokens = []
            char_to_word_offset = []
            prev_is_whitespace = True
            for c in paragraph_text:
                if is_whitespace(c):
                    prev_is_whitespace = True
                else:
                    if prev_is_whitespace:
                        doc_tokens.append(c)
                    else:
                        doc_tokens[-1] += c
                    prev_is_whitespace = False
                char_to_word_offset.append(len(doc_tokens) - 1)
            #-------paragraphs - End-------#
            qas_id = str(uuid.uuid1())
            example = SquadExample(
                qas_id=qas_id,
                question_text=question_text,
                doc_id = temp_list[i],
                doc_tokens=doc_tokens,
                orig_answer_text=orig_answer_text,
                start_position=start_position,
                end_position=end_position,
                is_impossible=is_impossible)
            eval_examples.append(example)
        #-------------documents - Start--------------#
    #-------------------------questions - End-----------------------------#
    if example_in_set_eval_examples == 1:
        print('len of eval_examples:%d' %len(eval_examples))
        for i, example in enumerate(eval_examples):
            print(i)
            print (example)
    return eval_examples



def main(_):
  tf.logging.set_verbosity(tf.logging.INFO)
  print(willy_check_code)
  
  bert_config = modeling.BertConfig.from_json_file(FLAGS.bert_config_file)

  validate_flags_or_throw(bert_config)

  tf.gfile.MakeDirs(FLAGS.output_dir)

  tokenizer = tokenization.FullTokenizer(
      vocab_file=FLAGS.vocab_file, do_lower_case=FLAGS.do_lower_case)

  tpu_cluster_resolver = None
  if FLAGS.use_tpu and FLAGS.tpu_name:
    tpu_cluster_resolver = tf.contrib.cluster_resolver.TPUClusterResolver(
        FLAGS.tpu_name, zone=FLAGS.tpu_zone, project=FLAGS.gcp_project)

  is_per_host = tf.contrib.tpu.InputPipelineConfig.PER_HOST_V2
  run_config = tf.contrib.tpu.RunConfig(
      cluster=tpu_cluster_resolver,
      master=FLAGS.master,
      model_dir=FLAGS.output_dir,
      save_checkpoints_steps=FLAGS.save_checkpoints_steps,
      tpu_config=tf.contrib.tpu.TPUConfig(
          iterations_per_loop=FLAGS.iterations_per_loop,
          num_shards=FLAGS.num_tpu_cores,
          per_host_input_for_training=is_per_host))

  train_examples = None
  num_train_steps = None
  num_warmup_steps = None
  
  ranker = None
  
  #------------------------do train(Start-(1))----------------------------#
  if FLAGS.do_train:
    train_examples = read_squad_examples(
        input_file=FLAGS.train_file, is_training=True)
    num_train_steps = int(
        len(train_examples) / FLAGS.train_batch_size * FLAGS.num_train_epochs)
    num_warmup_steps = int(num_train_steps * FLAGS.warmup_proportion)

    # Pre-shuffle the input to avoid having to make a very large shuffle
    # buffer in in the `input_fn`.
    rng = random.Random(12345)
    rng.shuffle(train_examples)
  #--------------------------do train(End-(1))----------------------------#

  model_fn = model_fn_builder(
      bert_config=bert_config,
      init_checkpoint=FLAGS.init_checkpoint,
      learning_rate=FLAGS.learning_rate,
      num_train_steps=num_train_steps,
      num_warmup_steps=num_warmup_steps,
      use_tpu=FLAGS.use_tpu,
      use_one_hot_embeddings=FLAGS.use_tpu)

  # If TPU is not available, this will fall back to normal Estimator on CPU
  # or GPU.
  estimator = tf.contrib.tpu.TPUEstimator(
      use_tpu=FLAGS.use_tpu,
      model_fn=model_fn,
      config=run_config,
      train_batch_size=FLAGS.train_batch_size,
      predict_batch_size=FLAGS.predict_batch_size)

  if FLAGS.do_train:
    # We write to a temporary file to avoid storing very large constant tensors
    # in memory.
    train_writer = FeatureWriter(
        filename=os.path.join(FLAGS.output_dir, "train.tf_record"),
        is_training=True)
    convert_examples_to_features(
        examples=train_examples,
        tokenizer=tokenizer,
        max_seq_length=FLAGS.max_seq_length,
        doc_stride=FLAGS.doc_stride,
        max_query_length=FLAGS.max_query_length,
        is_training=True,
        output_fn=train_writer.process_feature)
    train_writer.close()

    tf.logging.info("***** Running training *****")
    tf.logging.info("  Num orig examples = %d", len(train_examples))
    tf.logging.info("  Num split examples = %d", train_writer.num_features)
    tf.logging.info("  Batch size = %d", FLAGS.train_batch_size)
    tf.logging.info("  Num steps = %d", num_train_steps)
    del train_examples

    train_input_fn = input_fn_builder(
        input_file=train_writer.filename,
        seq_length=FLAGS.max_seq_length,
        is_training=True,
        drop_remainder=True)
    estimator.train(input_fn=train_input_fn, max_steps=num_train_steps)

  if FLAGS.do_predict:  
    
    
    #print('WillyTest(1)...do Set question:%s' %(FLAGS.question_type))
    #---------------------set question , changed by willy---------------------# 
    questions = list()
    '''
    if FLAGS.question_type is 'SQuAD':
        questions = read_squad_question(input_file=FLAGS.predict_file)
    elif FLAGS.question_type is 'one_question':
        questions.append(FLAGS.question)
    elif FLAGS.question_type is 'interactive':
        #TODO : interactive mode
        questions.append(FLAGS.question)
    '''
    if FLAGS.question_type == 'one_question':
        questions.append(FLAGS.question)
    elif FLAGS.question_type == 'questionTable':
        file = open(FLAGS.question_table, "r")
        for line in file.readlines():
            line = line.strip()
            # print line
            questions.append(line)
      
    
    #questions.append(FLAGS.question)
    
    #-------------------------------------------------------------------------#
    
    
    #--------------------set document , changed by willy--------------------# 
    if FLAGS.do_retriever:
        # Set Document
        #------------------------------------------------------

        print('WillyTest...do SQlite')
        DOC2IDX, docments = read_sqlite_documents(input_file=FLAGS.db_file)

        
        '''
        print('Show DOCIDX')
        print(DOC2IDX)
        for i, DOCID in enumerate(DOC2IDX) :
            print('ID:%d ,doc:%s' %(i,DOCID))
        '''
        #------------------------------------------------------
        
    else:
        # Set Document
        tf.logging.info("my document_type is %s",FLAGS.document_type)
        if FLAGS.document_type is 'Text':
            #TODO
            print('WillyTest...do Text')
            docments = read_text_documents(input_file=FLAGS.predict_file)
        
        elif FLAGS.document_type is 'SQuAD':
            #TODO
            print('WillyTest...do SQuAD')
            docments = read_squad_documents(input_file=FLAGS.predict_file)
        #else:
            #raise ValueError("Your document_type: %s is undefined or wrong, please reset it." %(FLAGS.document_type)) 

        #-------------------------------------------------------------------------#
        
    # define
    #---------------------------------------------------
    def append_feature(feature):
        eval_features.append(feature)
        eval_writer.process_feature(feature)
    # ---------------------------------------------------

    #print('WillyTest(2)...do Set eval_examples')
    eval_examples=set_eval_examples(questions,DOC2IDX)

    #print('WillyTest(2.1)...do FeatureWriter')
    eval_writer = FeatureWriter(
        filename=os.path.join(FLAGS.output_dir, "eval.tf_record"),
        is_training=False)
    eval_features = []

    #print('WillyTest(2.2)...do convert_examples_to_features')
    convert_examples_to_features(
        examples=eval_examples,
        tokenizer=tokenizer,
        max_seq_length=FLAGS.max_seq_length,
        doc_stride=FLAGS.doc_stride,
        max_query_length=FLAGS.max_query_length,
        is_training=False,
        output_fn=append_feature)
    eval_writer.close()    
    
    
    
    '''
    eval_examples = read_squad_examples(
        input_file=FLAGS.predict_file, is_training=False)
    
    
    print('WillyTest(3)...do FeatureWriter')
    eval_writer = FeatureWriter(
        filename=os.path.join(FLAGS.output_dir, "eval.tf_record"),
        is_training=False)
    eval_features = []
    print('WillyTest(4)...do convert_examples_to_features')
    convert_examples_to_features(
        examples=eval_examples,
        tokenizer=tokenizer,
        max_seq_length=FLAGS.max_seq_length,
        doc_stride=FLAGS.doc_stride,
        max_query_length=FLAGS.max_query_length,
        is_training=False,
        output_fn=append_feature)
    eval_writer.close()
    '''

    tf.logging.info("***** Running predictions *****")
    tf.logging.info("  Num orig examples = %d", len(eval_examples))
    tf.logging.info("  Num split examples = %d", len(eval_features))
    tf.logging.info("  Batch size = %d", FLAGS.predict_batch_size)


    print('WillyTest(5)...before redict_input_fn = input_fn_builder: eval_writer.filename=%s, FLAGS.max_seq_length=%d' %(eval_writer.filename,FLAGS.max_seq_length))
    predict_input_fn = input_fn_builder(
        input_file=eval_writer.filename,
        seq_length=FLAGS.max_seq_length,
        is_training=False,
        drop_remainder=False)

    # If running eval on the TPU, you will need to specify the number of
    # steps.
    all_results = []
    for result in estimator.predict(predict_input_fn, yield_single_examples=True):
        if len(all_results) % 1000 == 0:
            tf.logging.info("Processing example: %d" % (len(all_results)))
        unique_id = int(result["unique_ids"])
        start_logits = [float(x) for x in result["start_logits"].flat]
        end_logits = [float(x) for x in result["end_logits"].flat]
        all_results.append(RawResult(unique_id=unique_id,start_logits=start_logits,end_logits=end_logits))

    output_prediction_file = os.path.join(FLAGS.output_dir, "predictions.json")
    output_nbest_file = os.path.join(FLAGS.output_dir, "nbest_predictions.json")
    output_null_log_odds_file = os.path.join(FLAGS.output_dir, "null_odds.json")
    output_Aten_predict_file = os.path.join(FLAGS.output_dir, "Aten_predicts.json")

    print('WillyTest(8)...before write_predictions')  
    write_predictions(eval_examples, eval_features, all_results,
                      FLAGS.n_best_size, FLAGS.max_answer_length,
                      FLAGS.do_lower_case, output_prediction_file,
                      output_nbest_file, output_null_log_odds_file,
                      output_Aten_predict_file
                     )



if __name__ == "__main__":
  flags.mark_flag_as_required("vocab_file")
  flags.mark_flag_as_required("bert_config_file")
  flags.mark_flag_as_required("output_dir")
  tf.app.run()
